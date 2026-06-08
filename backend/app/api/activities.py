"""
Activities API: CRUD + statistics + charts for association activities.
"""
import io
from datetime import datetime, date, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from pydantic import BaseModel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.db.session import get_db
from app.models.user import User
from app.models.association_activity import AssociationActivity
from app.models.uploaded_file import UploadedFile
from app.api.auth import get_current_user
from app.api.dashboard import _get_latest_upload_ids

router = APIRouter()

ACTIVITY_LABELS = {
    "بازدید_علمی":          "بازدید علمی",
    "نشریه":                "نشریه",
    "سخنرانی_علمی":         "سخنرانی علمی",
    "همکاری_علمی":          "همکاری و مشارکت علمی",
    "کارگاه":               "برگزاری کارگاه",
    "نمایشگاه":             "برگزاری نمایشگاه",
    "دوره_کمک_آموزشی":      "دوره‌های کمک آموزشی",
    "همایش_علمی":           "برگزاری همایش علمی",
    "کرسی_اندیشه_ورزی":     "کرسی اندیشه‌ورزی",
    "مسابقه_علمی":          "مسابقه علمی",
}

ACTIVITY_COLORS = {
    "بازدید_علمی":      "#2174b8",
    "نشریه":            "#7fa343",
    "سخنرانی_علمی":     "#f59e0b",
    "همکاری_علمی":      "#8b5cf6",
    "کارگاه":           "#ec4899",
    "نمایشگاه":         "#06b6d4",
    "دوره_کمک_آموزشی":  "#f97316",
    "همایش_علمی":       "#10b981",
    "کرسی_اندیشه_ورزی": "#6366f1",
    "مسابقه_علمی":      "#ef4444",
}


def _get_latest_activity_upload_id(db: Session) -> Optional[int]:
    upload = db.query(UploadedFile).filter(
        UploadedFile.file_type == "activities",
        UploadedFile.status == "success",
    ).order_by(UploadedFile.parsed_at.desc()).first()
    return upload.id if upload else None


def _base_query(db: Session):
    uid = _get_latest_activity_upload_id(db)
    q = db.query(AssociationActivity)
    if uid:
        q = q.filter(AssociationActivity.upload_id == uid)
    return q


def _apply_period_filter(q, months: Optional[int]):
    if months:
        cutoff = datetime.now() - timedelta(days=months * 30)
        q = q.filter(
            AssociationActivity.year >= cutoff.year,
        )
    return q


# ── Endpoints ─────────────────────────────────────────────────────────────────

@router.get("")
def list_activities(
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    search: Optional[str] = None,
    activity_type: Optional[str] = None,
    months: Optional[int] = Query(None, description="Filter last N months"),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    q = _base_query(db)
    if search:
        q = q.filter(
            AssociationActivity.association_name.ilike(f"%{search}%") |
            AssociationActivity.title.ilike(f"%{search}%")
        )
    if activity_type:
        q = q.filter(AssociationActivity.activity_type == activity_type)
    if months:
        q = _apply_period_filter(q, months)

    total = q.count()
    items = q.order_by(AssociationActivity.year.desc(), AssociationActivity.month.desc(), AssociationActivity.id.desc()) \
             .offset((page - 1) * per_page).limit(per_page).all()

    return {
        "items": [_fmt(a) for a in items],
        "total": total,
        "page": page,
        "per_page": per_page,
        "total_pages": -(-total // per_page),
    }


@router.get("/stats")
def get_stats(
    months: Optional[int] = None,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    q = _base_query(db)
    if months:
        q = _apply_period_filter(q, months)

    total = q.count()

    # By type
    by_type = (
        q.with_entities(AssociationActivity.activity_type, func.count().label("cnt"))
         .group_by(AssociationActivity.activity_type)
         .all()
    )

    # By association (top 10)
    by_assoc = (
        q.with_entities(AssociationActivity.association_name, func.count().label("cnt"))
         .group_by(AssociationActivity.association_name)
         .order_by(func.count().desc())
         .limit(10)
         .all()
    )

    # By month (for trend chart)
    by_month = (
        q.with_entities(AssociationActivity.year, AssociationActivity.month, func.count().label("cnt"))
         .filter(AssociationActivity.year.isnot(None), AssociationActivity.month.isnot(None))
         .group_by(AssociationActivity.year, AssociationActivity.month)
         .order_by(AssociationActivity.year, AssociationActivity.month)
         .all()
    )

    # By quarter
    quarter_data: dict = {}
    for row in by_month:
        if row.year and row.month:
            q_label = f"{row.year}/Q{((row.month - 1) // 3) + 1}"
            quarter_data[q_label] = quarter_data.get(q_label, 0) + row.cnt

    return {
        "total": total,
        "by_type": [
            {
                "type": t,
                "label": ACTIVITY_LABELS.get(t, t),
                "color": ACTIVITY_COLORS.get(t, "#94a3b8"),
                "count": c,
            }
            for t, c in by_type
        ],
        "by_association": [
            {"name": name, "count": c} for name, c in by_assoc
        ],
        "monthly_trend": [
            {
                "label": f"{row.year}/{str(row.month).zfill(2)}",
                "year": row.year,
                "month": row.month,
                "count": row.cnt,
            }
            for row in by_month
        ],
        "quarterly_trend": [
            {"label": label, "count": cnt}
            for label, cnt in sorted(quarter_data.items())
        ],
        "activity_labels": ACTIVITY_LABELS,
        "activity_colors": ACTIVITY_COLORS,
    }


@router.get("/export")
def export_activities(
    months: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    q = _base_query(db)
    if months:
        q = _apply_period_filter(q, months)
    rows = q.order_by(AssociationActivity.year.desc(), AssociationActivity.month.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "فعالیت انجمن‌ها"
    ws.sheet_view.rightToLeft = True

    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor="1E40AF")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="E2E8F0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["نام انجمن", "نوع فعالیت", "عنوان", "تاریخ", "ترم", "سال تحصیلی",
               "تعداد شرکت‌کنندگان", "وضعیت", "توضیحات"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = border

    for ri, a in enumerate(rows, 2):
        vals = [
            a.association_name,
            ACTIVITY_LABELS.get(a.activity_type, a.activity_type),
            a.title or "",
            a.event_date or "",
            a.semester or "",
            a.academic_year or "",
            a.participants_count or "",
            a.status or "",
            a.description or "",
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
            cell.border = border

    for ci in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    period = f"{months}ماه" if months else "کامل"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=activities_{period}.xlsx"},
    )


def _fmt(a: AssociationActivity) -> dict:
    return {
        "id": a.id,
        "association_name": a.association_name,
        "activity_type": a.activity_type,
        "activity_label": ACTIVITY_LABELS.get(a.activity_type, a.activity_type),
        "title": a.title,
        "event_date": a.event_date,
        "semester": a.semester,
        "academic_year": a.academic_year,
        "month": a.month,
        "year": a.year,
        "participants_count": a.participants_count,
        "description": a.description,
        "status": a.status,
    }
