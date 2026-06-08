"""
Export API: Generate Excel and PDF reports.
Excel uses openpyxl with RTL, Persian headers, and org branding.
PDF uses ReportLab with Arabic reshaper for proper RTL rendering.
"""
import io
import json
from datetime import datetime
from typing import Optional
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.db.session import get_db
from app.models.user import User
from app.models.association import Association
from app.models.association_form import AssociationForm
from app.models.uploaded_file import UploadedFile
from app.api.auth import get_current_user
from app.core.config import settings

router = APIRouter()

# ── Style helpers ─────────────────────────────────────────────────────────────

PRIMARY_BLUE = "1E40AF"
LIGHT_BLUE = "EFF6FF"
SUCCESS_GREEN = "059669"
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor=PRIMARY_BLUE)
DATA_FONT = Font(name="Arial", size=10)
SUBHEADER_FILL = PatternFill("solid", fgColor=LIGHT_BLUE)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)
THIN = Side(style="thin", color="E2E8F0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _styled_header(ws, row: int, cols: list[str]) -> None:
    for col_idx, header in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER


def _get_latest_assoc_upload_id(db: Session) -> Optional[int]:
    upload = db.query(UploadedFile).filter(
        UploadedFile.file_type == 'associations',
        UploadedFile.status == 'success',
    ).order_by(UploadedFile.parsed_at.desc()).first()
    return upload.id if upload else None


def _get_latest_forms_upload_id(db: Session) -> Optional[int]:
    upload = db.query(UploadedFile).filter(
        UploadedFile.file_type == 'forms',
        UploadedFile.status == 'success',
    ).order_by(UploadedFile.parsed_at.desc()).first()
    return upload.id if upload else None


# ── Associations Excel ────────────────────────────────────────────────────────

@router.get("/associations.xlsx")
def export_associations(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    upload_id = _get_latest_assoc_upload_id(db)
    q = db.query(Association)
    if upload_id:
        q = q.filter(Association.upload_id == upload_id)
    records = q.order_by(Association.name).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "انجمن‌ها"
    ws.sheet_view.rightToLeft = True

    # Title row
    ws.merge_cells('A1:L1')
    ws['A1'] = f"{settings.ORG_NAME} — گزارش انجمن‌ها"
    ws['A1'].font = Font(name="Arial", bold=True, size=13, color=PRIMARY_BLUE)
    ws['A1'].alignment = CENTER

    ws.merge_cells('A2:L2')
    ws['A2'] = f"تاریخ تولید: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws['A2'].alignment = CENTER
    ws['A2'].font = Font(name="Arial", size=10, color="64748B")

    headers = ['#', 'نام انجمن', 'دبیر', 'شماره تماس', 'وضعیت فعالیت',
               'لوگو', 'هدر', 'ایمیل دانشجویی', 'کانال بله', 'تولید محتوا', 'نیاز به پیگیری', 'یادداشت']
    _styled_header(ws, 3, headers)

    for i, a in enumerate(records, start=1):
        row = i + 3
        missing = json.loads(a.missing_fields or '[]')
        ws.append([
            i, a.name, a.secretary_name or '', a.phone or '',
            a.activity_status, a.logo_status, a.header_status,
            a.student_email or '', a.channel_bale or '',
            a.content_production or '',
            'بله' if a.needs_follow_up else 'خیر',
            ', '.join(missing) if missing else '',
        ])
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = DATA_FONT
            cell.border = BORDER
            cell.alignment = RIGHT

    for col_idx, width in enumerate([5, 25, 18, 15, 15, 10, 10, 25, 12, 15, 14, 30], start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                              headers={"Content-Disposition": "attachment; filename*=UTF-8''%DA%AF%D8%B2%D8%A7%D8%B1%D8%B4-%D8%A7%D9%86%D8%AC%D9%85%D9%86%E2%80%8C%D9%87%D8%A7.xlsx"})


# ── Forms Excel ───────────────────────────────────────────────────────────────

@router.get("/forms.xlsx")
def export_forms(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    upload_id = _get_latest_forms_upload_id(db)
    q = db.query(AssociationForm)
    if upload_id:
        q = q.filter(AssociationForm.upload_id == upload_id)
    records = q.order_by(AssociationForm.association_name).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "فرم‌ها"
    ws.sheet_view.rightToLeft = True

    ws.merge_cells('A1:L1')
    ws['A1'] = f"{settings.ORG_NAME} — گزارش فرم‌ها"
    ws['A1'].font = Font(name="Arial", bold=True, size=13, color=PRIMARY_BLUE)
    ws['A1'].alignment = CENTER

    ws.merge_cells('A2:L2')
    ws['A2'] = f"تاریخ تولید: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws['A2'].alignment = CENTER

    headers = ['#', 'نام انجمن', 'نوع', 'عنوان فعالیت', 'تاریخ', 'وضعیت',
               'دبیر', 'مشاور', 'بازرس', 'رئیس دانشکده', 'کامل', 'پیگیری']
    _styled_header(ws, 3, headers)

    for i, f in enumerate(records, start=1):
        row = i + 3
        title = f.workshop_title or f.competition_title or ''
        ws.append([
            i, f.association_name, f.form_type, title,
            f.event_date or '', f.status or '',
            f.sig_secretary or '', f.sig_advisor or '',
            f.sig_inspector or '', f.sig_dean or '',
            'بله' if f.is_complete else 'خیر',
            'بله' if f.needs_follow_up else 'خیر',
        ])
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = DATA_FONT
            cell.border = BORDER
            cell.alignment = RIGHT

    for col_idx, width in enumerate([5, 25, 10, 30, 15, 15, 10, 10, 10, 15, 10, 10], start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                              headers={"Content-Disposition": "attachment; filename*=UTF-8''%DA%AF%D8%B2%D8%A7%D8%B1%D8%B4-%D9%81%D8%B1%D9%85%E2%80%8C%D9%87%D8%A7.xlsx"})


# ── Full report Excel ─────────────────────────────────────────────────────────

@router.get("/full-report.xlsx")
def export_full_report(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    assoc_id = _get_latest_assoc_upload_id(db)
    forms_id = _get_latest_forms_upload_id(db)

    wb = Workbook()
    wb.remove(wb.active)

    # ── Summary sheet ──
    ws_sum = wb.create_sheet("خلاصه")
    ws_sum.sheet_view.rightToLeft = True
    ws_sum['A1'] = settings.ORG_NAME
    ws_sum['A1'].font = Font(name="Arial", bold=True, size=14, color=PRIMARY_BLUE)
    ws_sum['A2'] = settings.APP_NAME
    ws_sum['A3'] = f"تاریخ تولید گزارش: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws_sum['A4'] = f"تهیه‌کننده: {current_user.full_name}"

    def aq():
        q = db.query(Association)
        if assoc_id:
            q = q.filter(Association.upload_id == assoc_id)
        return q

    def fq():
        q = db.query(AssociationForm)
        if forms_id:
            q = q.filter(AssociationForm.upload_id == forms_id)
        return q

    stats = [
        ("کل انجمن‌ها", aq().count()),
        ("انجمن‌های فعال", aq().filter(Association.activity_status == 'فعال').count()),
        ("انجمن‌های نیازمند پیگیری", aq().filter(Association.needs_follow_up == True).count()),
        ("کل فرم‌ها", fq().count()),
        ("فرم‌های کامل", fq().filter(AssociationForm.is_complete == True).count()),
        ("فرم‌های ناقص", fq().filter(AssociationForm.is_complete == False).count()),
    ]
    ws_sum['A6'] = "آمار خلاصه"
    ws_sum['A6'].font = Font(bold=True, size=12)
    for idx, (label, val) in enumerate(stats, start=7):
        ws_sum.cell(row=idx, column=1, value=label)
        ws_sum.cell(row=idx, column=2, value=val)
    ws_sum.column_dimensions['A'].width = 35
    ws_sum.column_dimensions['B'].width = 15

    # ── Associations sheet ──
    ws_a = wb.create_sheet("انجمن‌ها")
    ws_a.sheet_view.rightToLeft = True
    _styled_header(ws_a, 1, ['نام انجمن', 'دبیر', 'وضعیت فعالیت', 'لوگو', 'هدر', 'ایمیل', 'پیگیری'])
    for a in aq().all():
        ws_a.append([a.name, a.secretary_name or '', a.activity_status, a.logo_status, a.header_status, a.student_email or '', 'بله' if a.needs_follow_up else 'خیر'])

    # ── Forms sheet ──
    ws_f = wb.create_sheet("فرم‌ها")
    ws_f.sheet_view.rightToLeft = True
    _styled_header(ws_f, 1, ['انجمن', 'نوع', 'عنوان', 'تاریخ', 'کامل', 'پیگیری'])
    for f in fq().all():
        ws_f.append([f.association_name, f.form_type, f.workshop_title or f.competition_title or '', f.event_date or '', 'بله' if f.is_complete else 'خیر', 'بله' if f.needs_follow_up else 'خیر'])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                              headers={"Content-Disposition": "attachment; filename*=UTF-8''%DA%AF%D8%B2%D8%A7%D8%B1%D8%B4-%DA%A9%D8%A7%D9%85%D9%84.xlsx"})


# ── PDF Report ────────────────────────────────────────────────────────────────

@router.get("/dashboard.pdf")
def export_pdf(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    """Generate a branded Persian RTL PDF report using ReportLab."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import arabic_reshaper
    from bidi.algorithm import get_display

    def rtl(text: str) -> str:
        reshaped = arabic_reshaper.reshape(text)
        return get_display(reshaped)

    assoc_id = _get_latest_assoc_upload_id(db)
    forms_id = _get_latest_forms_upload_id(db)

    def aq():
        q = db.query(Association)
        if assoc_id:
            q = q.filter(Association.upload_id == assoc_id)
        return q

    def fq():
        q = db.query(AssociationForm)
        if forms_id:
            q = q.filter(AssociationForm.upload_id == forms_id)
        return q

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    story = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle('title', fontName='Helvetica-Bold', fontSize=16, alignment=1, spaceAfter=6)
    subtitle_style = ParagraphStyle('subtitle', fontName='Helvetica', fontSize=12, alignment=1, spaceAfter=4, textColor=colors.HexColor('#475569'))
    normal_style = ParagraphStyle('normal', fontName='Helvetica', fontSize=10, spaceAfter=4)
    section_style = ParagraphStyle('section', fontName='Helvetica-Bold', fontSize=12, spaceAfter=6, textColor=colors.HexColor('#1E40AF'))

    story.append(Paragraph(rtl(settings.ORG_NAME), title_style))
    story.append(Paragraph(rtl(settings.APP_NAME), subtitle_style))
    story.append(Paragraph(rtl(f"گزارش مدیریتی انجمن‌ها — {datetime.now().strftime('%Y-%m-%d')}"), normal_style))
    story.append(Spacer(1, 0.5*cm))

    story.append(Paragraph(rtl("آمار کلیدی"), section_style))

    stats_data = [
        [rtl("شاخص"), rtl("مقدار")],
        [rtl("کل انجمن‌ها"), str(aq().count())],
        [rtl("انجمن‌های فعال"), str(aq().filter(Association.activity_status == 'فعال').count())],
        [rtl("نیازمند پیگیری"), str(aq().filter(Association.needs_follow_up == True).count())],
        [rtl("کل فرم‌ها"), str(fq().count())],
        [rtl("فرم‌های کامل"), str(fq().filter(AssociationForm.is_complete == True).count())],
    ]

    t = Table(stats_data, colWidths=[12*cm, 4*cm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E40AF')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(t)

    doc.build(story)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf",
                              headers={"Content-Disposition": "attachment; filename*=UTF-8''%DA%AF%D8%B2%D8%A7%D8%B1%D8%B4-%D8%AF%D8%A7%D8%B4%D8%A8%D9%88%D8%B1%D8%AF.pdf"})
