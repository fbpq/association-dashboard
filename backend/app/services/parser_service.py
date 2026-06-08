"""
Excel Parser Service
Handles flexible parsing of Association and Form Excel files.
Tolerates missing columns, merged cells, empty rows, and Persian text variations.
"""
import json
import logging
import re
from pathlib import Path
from typing import Optional
import openpyxl
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models.association import Association
from app.models.association_form import AssociationForm
from app.models.association_activity import AssociationActivity
from app.models.uploaded_file import UploadedFile
from app.models.upload_log import UploadLog
from app.utils.persian import (
    normalize_persian, normalize_activity_status,
    extract_logo_header_status, normalize_presence,
    normalize_bool_value, is_cancelled,
)

logger = logging.getLogger(__name__)

# ── Expected column name mappings ────────────────────────────────────────────

ASSOC_COL_MAP = {
    'انجمن ها': 'name',
    'انجمن‌ها': 'name',
    'دبیر انجمن': 'secretary_name',
    'دبیر': 'secretary_name',
    'شماره تماس': 'phone',
    'شماره': 'phone',
    'توضیحات': 'notes',
    'فعالیت': 'activity_status',
    'لوگو و هدر(براساس سایت)': 'logo_header',
    'لوگو و هدر': 'logo_header',
    'ایمیل دانشجویی': 'student_email',
    'ایمیل': 'student_email',
    'جلسه نقشه راه': 'roadmap_session',
    'کانال بله': 'channel_bale',
    'کانال روبیکا': 'channel_rubika',
    'کانال ایگپ': 'channel_igap',
    'کانال ایتا': 'channel_ita',
    'تولید محتوا': 'content_production',
    'فرم مسابقه': 'form_competition',
    'فرم کارگاه': 'form_workshop',
    'وضعیت': 'status',
    'ثبت شدن فعالیت داخل سایت': 'site_activity_registered',
    'نقشه راه': 'roadmap',
    'جلسه برای انجمن هایی که تازه مجوز گرفتند': 'new_license_session',
}

ACTIVITY_COL_MAP = {
    'نام انجمن':                         'association_name',
    'انجمن':                             'association_name',
    'انجمن ها':                          'association_name',
    'انجمن‌ها':                          'association_name',
    'نوع فعالیت':                        'activity_type',
    'فعالیت':                            'activity_type',
    'عنوان فعالیت':                      'title',
    'عنوان':                             'title',
    'تاریخ برگزاری':                     'event_date',
    'تاریخ':                             'event_date',
    'ترم':                               'semester',
    'سال تحصیلی':                        'academic_year',
    'سال':                               'academic_year',
    'تعداد شرکت کنندگان':                'participants_count',
    'تعداد شرکت‌کنندگان':               'participants_count',
    'تعداد':                             'participants_count',
    'توضیحات':                           'description',
    'توضیح':                             'description',
    'وضعیت':                             'status',
}

# Map raw Persian activity type strings to canonical keys
ACTIVITY_TYPE_MAP = {
    'بازدید علمی':                    'بازدید_علمی',
    'بازدید':                         'بازدید_علمی',
    'نشریه':                          'نشریه',
    'سخنرانی علمی':                   'سخنرانی_علمی',
    'سخنرانی':                        'سخنرانی_علمی',
    'همکاری و مشارکت علمی':           'همکاری_علمی',
    'همکاری علمی':                    'همکاری_علمی',
    'همکاری':                         'همکاری_علمی',
    'برگزاری کارگاه':                 'کارگاه',
    'کارگاه':                         'کارگاه',
    'برگزاری نمایشگاه':               'نمایشگاه',
    'نمایشگاه':                       'نمایشگاه',
    'دوره های کمک آموزشی':           'دوره_کمک_آموزشی',
    'دوره‌های کمک آموزشی':            'دوره_کمک_آموزشی',
    'دوره کمک آموزشی':               'دوره_کمک_آموزشی',
    'برگزاری همایش علمی':             'همایش_علمی',
    'برگزاری همایش‌های علمی':         'همایش_علمی',
    'همایش علمی':                     'همایش_علمی',
    'همایش':                          'همایش_علمی',
    'کرسی اندیشه ورزی':              'کرسی_اندیشه_ورزی',
    'کرسی اندیشه‌ورزی':              'کرسی_اندیشه_ورزی',
    'کرسی':                           'کرسی_اندیشه_ورزی',
    'مسابقه علمی':                    'مسابقه_علمی',
    'مسابقه':                         'مسابقه_علمی',
}

FORM_COL_MAP = {
    'انجمن ها': 'association_name',
    'انجمن‌ها': 'association_name',
    'کارگاه': 'workshop_title',
    'مسابقات': 'competition_title',
    'تاریخ': 'event_date',
    'وضعیت': 'status',
    'امضا دبیر': 'sig_secretary',
    'امضا استاد مشاور': 'sig_advisor',
    'امضا بازرس': 'sig_inspector',
    'امضا رئیس دانشکده': 'sig_dean',
    'امضا رئیس اداره انجمن ها': 'sig_head_associations',
    'امضا مدیرکل': 'sig_director_general',
    'توضیح': 'description',
    'توضیحات': 'description',
}


def _add_log(db: Session, upload_id: int, level: str, message: str,
             row: Optional[int] = None, col: Optional[str] = None):
    db.add(UploadLog(upload_id=upload_id, level=level, message=message, row_number=row, column_name=col))


def _find_header_row(ws, col_map: dict) -> tuple[int, dict[str, int]]:
    """Find the header row and build column_name → col_index mapping."""
    for row_idx in range(1, min(10, ws.max_row + 1)):
        row_vals = [str(ws.cell(row=row_idx, column=c).value or '').strip() for c in range(1, ws.max_column + 1)]
        matched = 0
        col_index = {}
        for c_idx, val in enumerate(row_vals, start=1):
            normalized = normalize_persian(val) or val
            if normalized in col_map:
                field = col_map[normalized]
                if field not in col_index:  # first occurrence wins
                    col_index[field] = c_idx
                    matched += 1
        if matched >= 2:
            return row_idx, col_index
    return 1, {}


def _cell_val(ws, row: int, col: Optional[int]) -> Optional[str]:
    if col is None:
        return None
    val = ws.cell(row=row, column=col).value
    if val is None:
        return None
    return normalize_persian(str(val))


def parse_associations_file(db: Session, upload: UploadedFile) -> int:
    """Parse associations Excel file, store clean records, return record count."""
    file_path = Path(upload.file_path)
    wb = openpyxl.load_workbook(file_path, data_only=True)

    # Pick best sheet: most columns matching ASSOC_COL_MAP
    best_ws = None
    best_score = -1
    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        _, ci = _find_header_row(ws, ASSOC_COL_MAP)
        if len(ci) > best_score:
            best_score = len(ci)
            best_ws = ws

    ws = best_ws or wb.active
    header_row, col_index = _find_header_row(ws, ASSOC_COL_MAP)

    if 'name' not in col_index:
        _add_log(db, upload.id, 'warning', 'ستون «انجمن ها» پیدا نشد؛ پردازش با ستون‌های موجود ادامه می‌یابد.')

    records = 0
    last_valid_name: Optional[str] = None

    for row_idx in range(header_row + 1, ws.max_row + 1):
        # Skip fully empty rows
        row_values = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]
        if all(v is None or str(v).strip() == '' for v in row_values):
            continue

        name = _cell_val(ws, row_idx, col_index.get('name'))
        if name:
            last_valid_name = name
        else:
            name = last_valid_name

        if not name:
            _add_log(db, upload.id, 'warning', f'ردیف {row_idx}: نام انجمن خالی است و انجمن قبلی معتبری یافت نشد — ردیف نادیده گرفته شد.', row_idx)
            continue

        logo_header_raw = _cell_val(ws, row_idx, col_index.get('logo_header'))
        logo_status, header_status = extract_logo_header_status(logo_header_raw)

        activity_raw = _cell_val(ws, row_idx, col_index.get('activity_status'))
        activity_status = normalize_activity_status(activity_raw)

        # Determine missing fields
        missing: list[str] = []
        if logo_status in ('ندارد', 'نامشخص'):
            missing.append('logo_status')
        if header_status in ('ندارد', 'نامشخص'):
            missing.append('header_status')

        student_email = _cell_val(ws, row_idx, col_index.get('student_email'))
        if not student_email:
            missing.append('student_email')

        roadmap_session = _cell_val(ws, row_idx, col_index.get('roadmap_session'))
        if not roadmap_session or roadmap_session in ('برگزار نشده', 'خیر', 'ندارد'):
            missing.append('roadmap_session')

        channel_bale = _cell_val(ws, row_idx, col_index.get('channel_bale'))
        if not channel_bale or normalize_bool_value(channel_bale) is False:
            missing.append('channel_bale')

        site_reg = _cell_val(ws, row_idx, col_index.get('site_activity_registered'))
        if not site_reg or normalize_bool_value(site_reg) is False:
            missing.append('site_activity_registered')

        needs_follow_up = (
            activity_status in ('کم‌فعالیت', 'نامشخص')
            or len(missing) >= 2
        )

        assoc = Association(
            upload_id=upload.id,
            name=name,
            secretary_name=_cell_val(ws, row_idx, col_index.get('secretary_name')),
            phone=_cell_val(ws, row_idx, col_index.get('phone')),
            activity_status=activity_status,
            logo_status=logo_status,
            header_status=header_status,
            student_email=student_email,
            roadmap_session=roadmap_session,
            channel_bale=normalize_presence(channel_bale),
            channel_rubika=normalize_presence(_cell_val(ws, row_idx, col_index.get('channel_rubika'))),
            channel_igap=normalize_presence(_cell_val(ws, row_idx, col_index.get('channel_igap'))),
            channel_ita=normalize_presence(_cell_val(ws, row_idx, col_index.get('channel_ita'))),
            content_production=_cell_val(ws, row_idx, col_index.get('content_production')),
            form_competition=_cell_val(ws, row_idx, col_index.get('form_competition')),
            form_workshop=_cell_val(ws, row_idx, col_index.get('form_workshop')),
            status=_cell_val(ws, row_idx, col_index.get('status')),
            site_activity_registered=normalize_presence(site_reg),
            roadmap=_cell_val(ws, row_idx, col_index.get('roadmap')),
            new_license_session=_cell_val(ws, row_idx, col_index.get('new_license_session')),
            notes=_cell_val(ws, row_idx, col_index.get('notes')),
            needs_follow_up=needs_follow_up,
            missing_fields=json.dumps(missing, ensure_ascii=False),
        )
        db.add(assoc)
        records += 1

    db.flush()
    return records


def parse_forms_file(db: Session, upload: UploadedFile) -> int:
    """Parse forms Excel file, store clean records, return record count."""
    file_path = Path(upload.file_path)
    wb = openpyxl.load_workbook(file_path, data_only=True)

    best_ws = None
    best_score = -1
    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        _, ci = _find_header_row(ws, FORM_COL_MAP)
        if len(ci) > best_score:
            best_score = len(ci)
            best_ws = ws

    ws = best_ws or wb.active
    header_row, col_index = _find_header_row(ws, FORM_COL_MAP)

    records = 0
    last_assoc_name: Optional[str] = None

    for row_idx in range(header_row + 1, ws.max_row + 1):
        row_values = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]
        if all(v is None or str(v).strip() == '' for v in row_values):
            continue

        assoc_name = _cell_val(ws, row_idx, col_index.get('association_name'))
        if assoc_name:
            last_assoc_name = assoc_name
        else:
            assoc_name = last_assoc_name

        if not assoc_name:
            _add_log(db, upload.id, 'warning', f'ردیف {row_idx}: نام انجمن در فرم خالی است — ردیف نادیده گرفته شد.', row_idx)
            continue

        workshop = _cell_val(ws, row_idx, col_index.get('workshop_title'))
        competition = _cell_val(ws, row_idx, col_index.get('competition_title'))
        form_type = 'کارگاه' if workshop else ('مسابقه' if competition else 'نامشخص')

        sigs = {
            'sig_secretary':       _cell_val(ws, row_idx, col_index.get('sig_secretary')),
            'sig_advisor':         _cell_val(ws, row_idx, col_index.get('sig_advisor')),
            'sig_inspector':       _cell_val(ws, row_idx, col_index.get('sig_inspector')),
            'sig_dean':            _cell_val(ws, row_idx, col_index.get('sig_dean')),
            'sig_head_associations': _cell_val(ws, row_idx, col_index.get('sig_head_associations')),
            'sig_director_general':  _cell_val(ws, row_idx, col_index.get('sig_director_general')),
        }

        has_missing_sig = any(
            normalize_bool_value(v) is not True
            for v in sigs.values()
        )
        is_complete = not has_missing_sig

        desc = _cell_val(ws, row_idx, col_index.get('description'))
        cancelled = is_cancelled(desc)
        event_date = _cell_val(ws, row_idx, col_index.get('event_date'))

        needs_follow_up = has_missing_sig or not event_date or cancelled

        form = AssociationForm(
            upload_id=upload.id,
            association_name=assoc_name,
            workshop_title=workshop,
            competition_title=competition,
            event_date=event_date,
            status=_cell_val(ws, row_idx, col_index.get('status')),
            **{k: normalize_presence(v) for k, v in sigs.items()},
            description=desc,
            form_type=form_type,
            is_complete=is_complete,
            has_missing_signature=has_missing_sig,
            is_cancelled=cancelled,
            needs_follow_up=needs_follow_up,
        )
        db.add(form)
        records += 1

    db.flush()
    return records


def _parse_gregorian_date(date_str: Optional[str]) -> tuple[Optional[int], Optional[int]]:
    """Extract Gregorian year and month from a date string (handles Persian/Gregorian)."""
    if not date_str:
        return None, None
    # Try to extract 4-digit year and 1-2 digit month
    nums = re.findall(r'\d+', date_str)
    if not nums:
        return None, None
    # If first number looks like a year (1390-1420 Shamsi or 2015-2030 Gregorian)
    for n in nums:
        if 1390 <= int(n) <= 1420:  # Persian year → approximate Gregorian
            year = int(n) + 621
            month_candidates = [int(m) for m in nums if m != n and 1 <= int(m) <= 12]
            month = month_candidates[0] if month_candidates else None
            return year, month
        elif 2010 <= int(n) <= 2035:
            year = int(n)
            month_candidates = [int(m) for m in nums if m != n and 1 <= int(m) <= 12]
            month = month_candidates[0] if month_candidates else None
            return year, month
    return None, None


def parse_activities_file(db: Session, upload: UploadedFile) -> int:
    """Parse activities Excel file, store clean records, return record count."""
    file_path = Path(upload.file_path)
    wb = openpyxl.load_workbook(file_path, data_only=True)

    best_ws, best_score = None, -1
    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        _, ci = _find_header_row(ws, ACTIVITY_COL_MAP)
        if len(ci) > best_score:
            best_score, best_ws = len(ci), ws

    ws = best_ws or wb.active
    header_row, col_index = _find_header_row(ws, ACTIVITY_COL_MAP)

    records = 0
    last_assoc: Optional[str] = None

    for row_idx in range(header_row + 1, ws.max_row + 1):
        row_values = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]
        if all(v is None or str(v).strip() == '' for v in row_values):
            continue

        assoc_name = _cell_val(ws, row_idx, col_index.get('association_name'))
        if assoc_name:
            last_assoc = assoc_name
        else:
            assoc_name = last_assoc

        if not assoc_name:
            _add_log(db, upload.id, 'warning', f'ردیف {row_idx}: نام انجمن خالی — ردیف نادیده گرفته شد.', row_idx)
            continue

        raw_type = _cell_val(ws, row_idx, col_index.get('activity_type')) or ''
        canonical_type = ACTIVITY_TYPE_MAP.get(raw_type.strip(), raw_type.strip() or 'نامشخص')

        date_str = _cell_val(ws, row_idx, col_index.get('event_date'))
        year, month = _parse_gregorian_date(date_str)

        participants_raw = _cell_val(ws, row_idx, col_index.get('participants_count'))
        participants = None
        if participants_raw:
            nums = re.findall(r'\d+', participants_raw)
            if nums:
                participants = int(nums[0])

        activity = AssociationActivity(
            upload_id=upload.id,
            association_name=assoc_name,
            activity_type=canonical_type,
            title=_cell_val(ws, row_idx, col_index.get('title')),
            event_date=date_str,
            semester=_cell_val(ws, row_idx, col_index.get('semester')),
            academic_year=_cell_val(ws, row_idx, col_index.get('academic_year')),
            month=month,
            year=year,
            participants_count=participants,
            description=_cell_val(ws, row_idx, col_index.get('description')),
            status=_cell_val(ws, row_idx, col_index.get('status')) or 'برگزار شد',
        )
        db.add(activity)
        records += 1

    db.flush()
    return records


def process_upload(db: Session, upload: UploadedFile) -> None:
    """Main entry point: detect file type, parse, update upload status."""
    from datetime import datetime, timezone
    upload.status = 'processing'
    db.flush()

    warnings = 0
    try:
        _add_log(db, upload.id, 'info', f'پردازش فایل آغاز شد: {upload.original_filename}')

        if upload.file_type == 'associations':
            count = parse_associations_file(db, upload)
        elif upload.file_type == 'forms':
            count = parse_forms_file(db, upload)
        elif upload.file_type == 'activities':
            count = parse_activities_file(db, upload)
        else:
            # Auto-detect by trying all three
            try:
                count = parse_activities_file(db, upload)
                upload.file_type = 'activities'
            except Exception:
                try:
                    count = parse_associations_file(db, upload)
                    upload.file_type = 'associations'
                except Exception:
                    count = parse_forms_file(db, upload)
                    upload.file_type = 'forms'

        # Count warnings
        warnings = db.query(UploadLog).filter(
            UploadLog.upload_id == upload.id,
            UploadLog.level == 'warning',
        ).count()

        upload.status = 'success'
        upload.record_count = count
        upload.warnings_count = warnings
        upload.parsed_at = datetime.now(timezone.utc)
        _add_log(db, upload.id, 'info', f'پردازش با موفقیت پایان یافت — {count} رکورد استخراج شد، {warnings} هشدار ثبت شد.')
        db.commit()

    except Exception as e:
        db.rollback()
        upload.status = 'error'
        upload.error_message = str(e)
        db.add(UploadLog(upload_id=upload.id, level='error', message=f'خطای پردازش: {str(e)}'))
        db.commit()
        logger.error(f"Failed to process upload {upload.id}: {e}", exc_info=True)
